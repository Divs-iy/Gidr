# ✅ CORRECT STRUCTURE

# --- 1. IMPORTS (no duplicates) ---
import re
import os, json, shutil, logging
from database import get_db, init_db, InvoiceRecord, LineItem, User
#from huggingface_hub import User
import uvicorn
import pandas as pd
from datetime import datetime
from fastapi import FastAPI, File, UploadFile, HTTPException, Depends, Body, Query, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session, joinedload
from datetime import datetime, timedelta
from jose import JWTError, jwt
from fastapi.security import OAuth2PasswordBearer
import asyncio
from concurrent.futures import ThreadPoolExecutor
from pipeline.ocr import OCRProcessor
from pipeline.extractor import AIExtractor
from fastapi import Header
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import AuditReport, AuditItem
# --- PATCH START ---
import bcrypt
# Passlib looks for __about__, so we give it what it wants
def get_password_hash(password: str):
    # Convert string to bytes
    pwd_bytes = password.encode('utf-8')
    # Generate salt and hash the password
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(pwd_bytes, salt)
    return hashed.decode('utf-8')

def verify_password(plain_password: str, hashed_password: str):
    try:
        # Both the attempt and the stored hash must be encoded to bytes
        return bcrypt.checkpw(
            plain_password.encode('utf-8'), 
            hashed_password.encode('utf-8')
        )
    except Exception:
        return False
# --- PATCH END ---

# --- 2. APP SETUP ---
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="Gidr API")
executor = ThreadPoolExecutor(max_workers=2)
EXPORT_DIR = "data/exports"
if not os.path.exists(EXPORT_DIR):
    os.makedirs(EXPORT_DIR, exist_ok=True)
# app.mount("/data/raw", StaticFiles(directory="data/raw"), name="raw_files")
app.mount("/downloads", StaticFiles(directory="data/exports"), name="downloads")
# main.py — add this line after the other app.mount lines
os.makedirs("data/raw_archive", exist_ok=True)
app.mount("/originals", StaticFiles(directory="data/raw_archive"), name="originals")
app.add_middleware(CORSMiddleware, allow_credentials=True, allow_methods=["*"], allow_origins=["*"], allow_headers=["*"])

ocr_engine = OCRProcessor()
extractor_engine = AIExtractor()
init_db()
# Security Constants
SECRET_KEY = "GIDR_SUPER_SECRET_KEY_2026" # Keep this safe!
ALGORITHM = "HS256"
# pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="login")

# Helper Functions
def get_password_hash(password: str) -> str:
    pwd_bytes = password.encode('utf-8')
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(pwd_bytes, salt)
    return hashed.decode('utf-8')

def verify_password(plain_password: str, hashed_password: str) -> bool:
    try:
        return bcrypt.checkpw(
            plain_password.encode('utf-8'),
            hashed_password.encode('utf-8')
        )
    except Exception:
        return False


def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(hours=24)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


@app.post("/register")
async def register(
    email: str = Form(...), 
    password: str = Form(...), 
    db: Session = Depends(get_db)
):
    # 1. Check if user already exists
    existing_user = db.query(User).filter(User.email == email).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")

    # 2. Hash and save (using the new bcrypt logic)
    hashed = get_password_hash(password)
    new_user = User(email=email, hashed_password=hashed)
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    return {"msg": "Registration successful", "user_id": new_user.id}

@app.post("/login")
async def login(email: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == email).first()
    if not user or not verify_password(password, user.hashed_password):
        raise HTTPException(status_code=400, detail="Invalid credentials")
    
    token = create_access_token(data={"sub": user.email, "id": user.id})
    return {"access_token": token, "token_type": "bearer", "user_name": user.email}

# --- 3. HELPERS ---
def save_processed_json(filename: str, data: dict):
    """Saves extracted data so /compare can load it without re-OCR."""
    processed_dir = "data/processed"
    os.makedirs(processed_dir, exist_ok=True)
    path = os.path.join(processed_dir, f"{filename}.json")
    with open(path, "w") as f:
        json.dump(data, f)
    return path

# --- 4. ROUTES ---
async def get_current_user(authorization: str = Header(None)):
    if not authorization:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    try:
        # Assumes format "Bearer <token>"
        token = authorization.split(" ")[1]
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: int = payload.get("id")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        return user_id
    except Exception:
        raise HTTPException(status_code=401, detail="Could not validate credentials")
@app.post("/upload")
async def upload_invoice(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user_id: int = Depends(get_current_user),
    source: str = Query(default="extraction")
):
    logger.info(f"Received file: {file.filename}")
    if not file.filename.lower().endswith(('.png', '.jpg', '.jpeg', '.pdf')):
        raise HTTPException(status_code=400, detail="Invalid file type.")

    temp_path = f"data/raw/{file.filename}"
    os.makedirs("data/raw", exist_ok=True)

    with open(temp_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    try:
        loop = asyncio.get_event_loop()

        def run_pipeline():
            ocr_results = ocr_engine.run_ocr(temp_path)
            return extractor_engine.extract_with_groq(ocr_results)

        extracted = await loop.run_in_executor(executor, run_pipeline)

        base_name = os.path.splitext(file.filename)[0]
        save_processed_json(base_name, extracted)

        rows = []
        for item in extracted.get("line_items", []):
            rows.append({
                "Invoice Number": extracted.get("invoice_number"),
                "Vendor": extracted.get("vendor_name"),
                "Date": extracted.get("date"),
                "Description": item.get("description"),
                "Quantity": item.get("quantity"),
                "Unit Price": item.get("unit_price"),
                "Line Amount": item.get("amount"),
                "Total Bill": extracted.get("total_amount"),
                "Terms & Conditions": extracted.get("terms_and_conditions", ""),
            })
        if not rows:
            rows.append({"Note": "No line items found", **extracted})
        BASE_URL = os.getenv("RENDER_EXTERNAL_URL", "http://localhost:8000")
        file_name = f"invoice_{extracted.get('invoice_number', 'unknown')}_{datetime.now().strftime('%H%M%S')}.xlsx"
        web_export_path = os.path.join("data", "exports", file_name)
        pd.DataFrame(rows).to_excel(web_export_path, index=False)
        download_url = f"{BASE_URL}/downloads/{file_name}"

        # ✅ Only save to DB if real extraction — quotes are temp files only
        if source == "extraction":
            new_record = InvoiceRecord(
                vendor_name=extracted.get("vendor_name", "Unknown"),
                total_amount=extracted.get("total_amount", 0.0),
                invoice_number=extracted.get("invoice_number", "N/A"),
                date=extracted.get("date"),
                excel_link=download_url,
                confidence_score=extracted.get("confidence_score", None),
                user_id=current_user_id,
                original_filename=file.filename,
                terms_and_conditions=extracted.get("terms_and_conditions", ""),
            )
            db.add(new_record)
            db.flush()

            for item in extracted.get("line_items", []):
                db.add(LineItem(
                    invoice_id=new_record.id,
                    description=item.get("description"),
                    unit=item.get("unit"),
                    quantity=item.get("quantity"),
                    unit_price=item.get("unit_price"),
                    amount=item.get("amount")
                ))
            db.commit()

        # ✅ Always return regardless of source
        return {
            "status": "success",
            "filename": file.filename,
            "saved_as": base_name,
            "download_url": download_url,
            "extracted_data": extracted,
            # "record_id": new_record.id if new_record else None
        }

    except Exception as e:
        db.rollback()
        logger.error(f"Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        if os.path.exists(temp_path):
            # ✅ Only archive originals for real extractions
            if source == "extraction":
                archive_path = f"data/raw_archive/{file.filename}"
                os.makedirs("data/raw_archive", exist_ok=True)
                shutil.copy(temp_path, archive_path)
            os.remove(temp_path)

@app.post("/compare")
async def compare_docs(
    invoice: UploadFile = File(...),
    quote_filename: str = Query(...)
):
    quote_path = f"data/processed/{quote_filename}.json"
    if not os.path.exists(quote_path):
        raise HTTPException(status_code=404, detail="Quote not found. Upload quote first.")

    with open(quote_path, "r") as f:
        quote_data = json.load(f)

    temp_invoice = f"data/raw/{invoice.filename}"
    os.makedirs("data/raw", exist_ok=True)

    try:
        with open(temp_invoice, "wb") as buffer:
            shutil.copyfileobj(invoice.file, buffer)

        invoice_base = os.path.splitext(invoice.filename)[0]
        invoice_json_path = f"data/processed/{invoice_base}.json"

        if os.path.exists(invoice_json_path):
            with open(invoice_json_path, "r") as f:
                invoice_data = json.load(f)
            logger.info(f"Using cached invoice JSON: {invoice_json_path}")
        else:
            loop = asyncio.get_event_loop()
            def run_compare_pipeline():
                ocr_result = ocr_engine.run_ocr(temp_invoice)
                return extractor_engine.extract_with_groq(ocr_result)
            invoice_data = await loop.run_in_executor(executor, run_compare_pipeline)
            save_processed_json(invoice_base, invoice_data)

        discrepancies = []

        # Total amount check
        q_amt = round(float(quote_data.get('total_amount') or 0), 2)
        i_amt = round(float(invoice_data.get('total_amount') or 0), 2)
        if abs(q_amt - i_amt) > 0.01:
            discrepancies.append({
                "field": "Total Amount",
                "quote": q_amt,
                "invoice": i_amt,
                "diff": round(i_amt - q_amt, 2)
            })

        # Line item check — match by rounded amount only
        q_items_list = quote_data.get('line_items', [])
        i_items_list = invoice_data.get('line_items', [])

        # Build invoice amount lookup
        i_amounts = {}
        for item in i_items_list:
            key = round(float(item.get('amount') or 0), 2)
            i_amounts[key] = item

        for q_item in q_items_list:
            q_val = round(float(q_item.get('amount') or 0), 2)
            q_desc = q_item.get('description', 'Item')[:50]
            if q_val == 0:
                continue
            if q_val not in i_amounts:
                discrepancies.append({
                    "field": f"'{q_desc}'",
                    "quote": q_val,
                    "invoice": "MISSING",
                    "diff": q_val
                })
            # amount matched — no discrepancy, skip

        return {
            "status": "MISMATCH" if discrepancies else "MATCH",
            "discrepancies": discrepancies,
            "invoice_data": invoice_data,
            "quote_data": quote_data
        }

    finally:
        if os.path.exists(temp_invoice):
            os.remove(temp_invoice)

@app.put("/invoices/{invoice_id}")
async def update_invoice(invoice_id: int, updated_data: dict = Body(...), db: Session = Depends(get_db), current_user_id: int = Depends(get_current_user)):
    invoice = db.query(InvoiceRecord).filter(InvoiceRecord.id == invoice_id, InvoiceRecord.user_id == current_user_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    invoice.vendor_name = updated_data.get("vendor_name", invoice.vendor_name)
    invoice.total_amount = updated_data.get("total_amount", invoice.total_amount)
    invoice.date = updated_data.get("date", invoice.date)
    db.commit()
    return {"status": "success", "message": "Updated successfully"}


@app.delete("/invoices/{invoice_id}")
async def delete_invoice(invoice_id: int, db: Session = Depends(get_db), current_user_id: int = Depends(get_current_user)):
    invoice = db.query(InvoiceRecord).filter(InvoiceRecord.id == invoice_id, InvoiceRecord.user_id == current_user_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    db.delete(invoice)
    db.commit()
    return {"status": "success", "message": "Deleted successfully"}

@app.get("/history")
async def get_history(
    db: Session = Depends(get_db),
    current_user_id: int = Depends(get_current_user)
):
    invoices = db.query(InvoiceRecord).options(
        joinedload(InvoiceRecord.items)
    ).filter(
        InvoiceRecord.user_id == current_user_id
    ).order_by(InvoiceRecord.id.desc()).all()

    # ✅ Manually serialize — fixes history page showing nothing
    return [
        {
            "id": inv.id,
            "vendor_name": inv.vendor_name,
            "total_amount": inv.total_amount,
            "date": inv.date,
            "confidence_score": inv.confidence_score,
            # "filename": inv.invoice_number,
            "filename": inv.original_filename or inv.invoice_number,
            "excel_link": inv.excel_link,
            "items": [
    {
        "id": item.id,
        "description": item.description,
        "unit": item.unit,
        "quantity": item.quantity,
        "amount": item.amount
    }
    for item in inv.items
]
        }
        for inv in invoices
    ]


@app.get("/export-all")
async def export_all(db: Session = Depends(get_db), current_user_id: int = Depends(get_current_user)):
    invoices = db.query(InvoiceRecord).filter(InvoiceRecord.user_id == current_user_id).all()
    all_data = []
    for inv in invoices:
        for item in inv.items:
            all_data.append({
                "Invoice #": inv.invoice_number,
                "Vendor": inv.vendor_name,
                "Item": item.description,
                "Amount": item.amount,
                "Total Bill": inv.total_amount,
                "Date": inv.date
            })
    master_path = os.path.expanduser("~/Desktop/Gidr_Exports/Master_Report.xlsx")
    pd.DataFrame(all_data).to_excel(master_path, index=False)
    return {"status": "success", "path": master_path}

def generate_audit_reason(status: str, quoted: float, invoiced: float, desc: str) -> str:
    if status == "MATCH":
        return "Price matches quotation exactly."
    elif status == "PRICE MISMATCH":
        diff = round(invoiced - quoted, 2)
        if diff > 0:
            return f"Invoiced price is ₹{diff} higher than quoted price."
        else:
            return f"Invoiced price is ₹{abs(diff)} lower than quoted price."
    elif status == "NOT IN QUOTE":
        return "Item was never approved in the initial quotation."
    elif status == "MISSING IN INVOICE":
        return "Item present in quote but missing from invoice."
    return ""

@app.post("/audit/save")
async def save_audit(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user_id: int = Depends(get_current_user)
):

    # Delete previous audit for same quote+invoice combo
    existing = db.query(AuditReport).filter(
        AuditReport.user_id == current_user_id,
        AuditReport.quote_filename == payload.get("quote_filename"),
        AuditReport.invoice_filename == payload.get("invoice_filename")
    ).first()
    if existing:
        db.delete(existing)
        db.flush()

    report = AuditReport(
        user_id=current_user_id,
        quote_filename=payload.get("quote_filename"),
        invoice_filename=payload.get("invoice_filename")
    )
    db.add(report)
    db.flush()

    for item in payload.get("items", []):
        db.add(AuditItem(
            report_id=report.id,
            description=item.get("description"),
            quoted_price=item.get("quoted_price", 0),
            invoiced_price=item.get("invoiced_price", 0),
            status=item.get("status"),
            reason=item.get("reason"),
            action=item.get("action", "FLAGGED"),
            comment=item.get("comment", "")
        ))
    db.commit()
    return {"status": "success", "report_id": report.id}


@app.get("/audit/report/{report_id}")
async def download_audit_report(
    report_id: int,
    db: Session = Depends(get_db),
    current_user_id: int = Depends(get_current_user)
):
    from database import AuditReport, AuditItem
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    report = db.query(AuditReport).filter(
        AuditReport.id == report_id,
        AuditReport.user_id == current_user_id
    ).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Report"

    # Styles
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1F2D3D")
    red_fill = PatternFill("solid", start_color="FFCCCC")
    green_fill = PatternFill("solid", start_color="CCFFCC")
    yellow_fill = PatternFill("solid", start_color="FFFACC")
    grey_fill = PatternFill("solid", start_color="F0F0F0")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Title row
    ws.merge_cells("A1:G1")
    ws["A1"] = "GIDR — VENDOR AUDIT REPORT"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", start_color="1F2D3D")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    # Meta row
    ws.merge_cells("A2:G2")
    ws["A2"] = f"Quote: {report.quote_filename}   |   Invoice: {report.invoice_filename}   |   Generated: {report.created_at.strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(name="Arial", size=9, italic=True)
    ws["A2"].fill = grey_fill
    ws["A2"].alignment = left
    ws.row_dimensions[2].height = 18

    # Headers
    headers = ["Item Description", "Quoted Price (₹)", "Invoiced Price (₹)", "Status", "Reason", "Action", "Accountant Note"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[3].height = 22

    # Data rows
    for row_idx, item in enumerate(report.items, start=4):
        values = [
            item.description,
            item.quoted_price,
            item.invoiced_price,
            item.status,
            item.reason,
            item.action,
            item.comment or ""
        ]
        # Pick row fill based on action + status
        if item.status == "MATCH":
            row_fill = green_fill
        elif item.action == "APPROVED":
            row_fill = yellow_fill
        else:
            row_fill = red_fill

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.border = thin
            cell.alignment = center if col != 1 else left
            if col in [2, 3]:
                cell.number_format = '#,##0.00'

        ws.row_dimensions[row_idx].height = 20

    # Summary rows
    items = report.items
    total_rows = len(items)
    flagged = sum(1 for i in items if i.action == "FLAGGED" and i.status != "MATCH")
    approved = sum(1 for i in items if i.action == "APPROVED")
    matched = sum(1 for i in items if i.status == "MATCH")
    flagged_amount = sum(
        abs((i.invoiced_price or 0) - (i.quoted_price or 0))
        for i in items if i.action == "FLAGGED" and i.status != "MATCH"
    )

    summary_row = total_rows + 5
    ws.merge_cells(f"A{summary_row}:G{summary_row}")
    ws[f"A{summary_row}"] = "SUMMARY"
    ws[f"A{summary_row}"].font = Font(bold=True, name="Arial", color="FFFFFF")
    ws[f"A{summary_row}"].fill = PatternFill("solid", start_color="1F2D3D")
    ws[f"A{summary_row}"].alignment = center

    summary_data = [
        ("Total Items Reviewed", total_rows),
        ("Matched Items", matched),
        ("Flagged Items", flagged),
        ("Approved Exceptions", approved),
        ("Total Disputed Amount (₹)", f"₹{flagged_amount:,.2f}"),
    ]
    for i, (label, val) in enumerate(summary_data):
        r = summary_row + 1 + i
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, name="Arial")
        ws.cell(row=r, column=2, value=val)
        ws.cell(row=r, column=1).border = thin
        ws.cell(row=r, column=2).border = thin

    # Column widths
    col_widths = [40, 18, 18, 20, 45, 14, 40]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Save
    os.makedirs("data/exports", exist_ok=True)
    path = f"data/exports/audit_report_{report_id}.xlsx"
    wb.save(path)

    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"Gidr_Audit_Report_{report_id}.xlsx"
    )


@app.get("/audit/list")
async def list_audits(
    db: Session = Depends(get_db),
    current_user_id: int = Depends(get_current_user)
):

    reports = db.query(AuditReport).filter(
        AuditReport.user_id == current_user_id
    ).order_by(AuditReport.id.desc()).all()
    return [
        {
            "id": r.id,
            "quote_filename": r.quote_filename,
            "invoice_filename": r.invoice_filename,
            "created_at": r.created_at.isoformat(),
            "item_count": len(r.items)
        }
        for r in reports
    ]

@app.post("/upload/smart")
async def upload_smart(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user_id: int = Depends(get_current_user)
):
    logger.info(f"Smart upload: {file.filename}")
    if not file.filename.lower().endswith(('.png', '.jpg', '.jpeg', '.pdf')):
        raise HTTPException(status_code=400, detail="Invalid file type.")

    temp_path = f"data/raw/{file.filename}"
    os.makedirs("data/raw", exist_ok=True)

    with open(temp_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    try:
        loop = asyncio.get_event_loop()

        def run_smart_pipeline():
            ocr_results = ocr_engine.run_ocr(temp_path)
            return extractor_engine.extract_smart(ocr_results)

        extracted = await loop.run_in_executor(executor, run_smart_pipeline)

        # Build Excel — two sheets
        os.makedirs("data/exports", exist_ok=True)
        os.makedirs("data/raw", exist_ok=True)
        doc_type = extracted.get("document_type", "DOCUMENT")
        safe_type = re.sub(r'[^a-z0-9_]', '', doc_type.lower().replace(' ', '_').replace('/', '_'))
        file_name = f"smart_{safe_type}_{datetime.now().strftime('%H%M%S')}.xlsx"
        export_path = os.path.join("data", "exports", file_name)
        with pd.ExcelWriter(export_path, engine="openpyxl") as writer:

            # Sheet 1: Summary — key fields + parties + dates + financials
            summary_rows = []

            # Parties
            parties = extracted.get("parties", {})
            for role, party in parties.items():
                if party.get("name"):
                    summary_rows.append({
                        "Category": "Party",
                        "Field": party.get("role") or role,
                        "Value": f"{party.get('name')} — {party.get('location', '')}"
                    })

            # Dates
            dates = extracted.get("dates", {})
            for label, val in dates.items():
                if val and label != "other_dates":
                    summary_rows.append({
                        "Category": "Date",
                        "Field": label.replace("_", " ").title(),
                        "Value": val
                    })
            for d in dates.get("other_dates", []):
                if d:
                    val = " | ".join(f"{k}: {v}" for k, v in d.items()) if isinstance(d, dict) else str(d)
                    summary_rows.append({"Category": "Date", "Field": "Other", "Value": val})

            # Key fields
            for kf in extracted.get("key_fields", []):
                summary_rows.append({
                    "Category": "Key Info",
                    "Field": kf.get("label"),
                    "Value": kf.get("value")
                })

            # Financials
            fin = extracted.get("financials", {})
            if fin.get("total_amount"):
                summary_rows.append({
                    "Category": "Financial",
                    "Field": "Total Amount",
                    "Value": f"{fin.get('currency', '')} {fin.get('total_amount', 0)}"
                })
            for ps in fin.get("payment_schedule", []):
                if isinstance(ps, dict):
                    val = " | ".join(f"{k}: {v}" for k, v in ps.items())
                else:
                    val = str(ps)
                summary_rows.append({
                "Category": "Payment Schedule",
                "Field": "",
                "Value": val
                })

            # Misc
            if extracted.get("cancellation_policy"):
                summary_rows.append({
                    "Category": "Cancellation",
                    "Field": "Policy",
                    "Value": extracted.get("cancellation_policy")
                })
            if extracted.get("governing_law"):
                summary_rows.append({
                    "Category": "Legal",
                    "Field": "Governing Law",
                    "Value": extracted.get("governing_law")
                })

            pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)

            # Sheet 2: Fee Breakdown
            fee_rows = fin.get("fee_breakdown", [])
            if fee_rows:
                pd.DataFrame(fee_rows).to_excel(writer, sheet_name="Fee Breakdown", index=False)

            # Sheet 3: Sections & T&C
            tc_rows = []
            for section in extracted.get("sections", []):
                tc_rows.append({
                    "Section": f"{section.get('section_number', '')} {section.get('title', '')}".strip(),
                    "Summary": section.get("summary", ""),
                    "Key Points": " | ".join(section.get("key_points", []))
                })
            for tc in extracted.get("terms_and_conditions", []):
                tc_rows.append({
                    "Section": tc.get("category", "Terms"),
                    "Summary": tc.get("condition", ""),
                    "Key Points": ""
                })
            if tc_rows:
                pd.DataFrame(tc_rows).to_excel(writer, sheet_name="Sections & Terms", index=False)

        BASE_URL = os.getenv("RENDER_EXTERNAL_URL", "http://localhost:8000")
        download_url = f"{BASE_URL}/downloads/{file_name}"

        return {
            "status": "success",
            "document_type": doc_type,
            "download_url": download_url,
            "summary": {
                "document_type": doc_type,
                "title": extracted.get("title", file.filename),
                "confidence": extracted.get("confidence", 0.0),
                "key_fields": extracted.get("key_fields", []),
                "sections_found": [
                    f"{s.get('section_number', '')} {s.get('title', '')}".strip()
                    for s in extracted.get("sections", [])
                ]
            }
        }

    except Exception as e:
        logger.error(f"Smart upload error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)
if __name__ == "__main__":
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=False)